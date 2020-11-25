from django.shortcuts import render,redirect
import logging
from .models import Employee,ExcelFiles
from .forms import EmployeeForms,SuperVisorForm,ExcelUploadForm
from users.models import User
import openpyxl
from utils.date_format import getAge
from utils.excel_utils import checkduplicateEmails
from .task import createEmployee
from datetime import date
# Create your views here.
def employeeListView(request):
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    else:
        employees = Employee.objects.all()
        return render(request,'employees.html',{'employees':employees})

def addEmployee(request):
    form = EmployeeForms()
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    else:
        if request.method == 'GET':
            return render(request,'add_employee.html',{'form':form})
        else:
            form = EmployeeForms(data=request.POST)
            if form.is_valid():
                user = User.objects.create_user(request.POST['email'],password=request.POST['password'])
                employee = Employee.objects.create(first_name = request.POST['first_name'],last_name = request.POST['first_name'],date_of_birth = request.POST['date_of_birth'],
                date_of_employment = request.POST['date_of_employment'],position = request.POST['position'],salary = request.POST['salary'],user_instance = user
                 )
                
                new_emp = Employee.objects.filter(pk  =employee.pk ).first()
                new_emp.age = getAge(new_emp.date_of_birth)
                new_emp.save()
                return redirect('/employees/congrats')
            else:
                return render(request,'add_employee.html',{'form':form})

def supervisorsView(request):
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    supervisors = Employee.objects.filter(is_supervisor = True)
    return render(request,'supervisors.html',{'supervisors':supervisors})

def congratsPageView(request):
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    return render(request,'employee_congrats.html')

def addSupervisorView(request):
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    employess = Employee.objects.filter(is_supervisor = False)
    return render(request,'add_supervisor.html',{'employees':employess})

def makeSupervisorView(request,pk):
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    try:
        employee = Employee.objects.get(pk = pk)
        employee.is_supervisor = True
        employee.save()
        return redirect('/employees/supervisors')
    except Employee.DoesNotExist:
        return redirect('/employees/supervisors/add')

def assignSupervisor(request):
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    form = SuperVisorForm(request.POST or None)
    if request.method == 'GET':
        return render(request,'assign_supervisro.html',{'form':form})
    else:
        form = SuperVisorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/employees/assign-supervisor/congrats')
        else:
            return render(request,'assign_supervisro.html',{'form':form})

def assignCongratsPageView(request):
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    return render(request,'assign_supervisor_congrats.html')
def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

def uploadExcelPageView(request):
    form = ExcelUploadForm(request.POST or None,request.FILES or None)
    user = request.user
    if not user.is_authenticated:
        return redirect('/auth/login')
    if request.method == "GET":
        return render(request,'upload_excel.html',{'form':form})
    else:
        list_of_emails = list()
        if form.is_valid():
            exscel = form.save()
            excel_path =ExcelFiles.objects.get(activated = True)
            excel_path.activated = False
            excel_path.save()
            wb  =openpyxl.load_workbook(excel_path.file)
            
            sheet = wb.worksheets[0]
            for row in list(iter_rows(sheet))[1:]:
                
               #Compare current email with existing if to see if its duplocated
                isDuplicate =  checkduplicateEmails(list_of_emails,row[2])
                if isDuplicate:
                    logging.ERROR('Duplicate email')
                    continue
                else:
                    createEmployee(row)
                list_of_emails.append(row[2])
            excel_path.delete()
            return redirect('/employees')
        else:
            return render(request,'upload_excel.html',{'form':form})